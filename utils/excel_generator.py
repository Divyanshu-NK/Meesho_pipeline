# utils/excel_generator.py
import openpyxl
from openpyxl.styles import PatternFill
import io

def generate_excel(rows_data):
    """Generate Excel file for Meesho upload"""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Product Name","Variation","Meesho Price","MRP","GST %",
               "Image Link 1","Image Link 2","Image Link 3","Image Link 4","Image Link 5",
               "Seller SKU","Brand Name","Product ID","Description","HSN Code","Weight (g)","Keywords"]
    
    # Add headers with styling
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="FF0000" if c<=10 else "00FF00")
    
    # Add data rows
    for r, row in enumerate(rows_data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer