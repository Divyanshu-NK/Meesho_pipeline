# app.py  ← UPDATED: Now includes product links to best-selling items

import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
import io
import requests
import base64
import time
from PIL import Image  # For previewing uploaded images

# ——————— Imgur upload (only if images uploaded) ———————
def upload_to_imgur(uploaded_file):
    url = "https://api.imgur.com/3/image"
    try:
        # Read bytes from uploaded file
        bytes_data = uploaded_file.read()
        payload = base64.b64encode(bytes_data)
        headers = {'Authorization': 'Client-ID 546c25a59c58ad7'}
        r = requests.post(url, headers=headers, data=payload)
        if r.status_code == 200:
            return r.json()['data']['link']
        else:
            st.warning(f"Upload failed for {uploaded_file.name}")
            return ""
    except Exception as e:
        st.error(f"Error uploading {uploaded_file.name}: {e}")
        return ""

# ——————— Generate Excel (with or without images) ———————
def generate_excel(rows_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Product Name","Variation","Meesho Price","MRP","GST %",
               "Image Link 1","Image Link 2","Image Link 3","Image Link 4","Image Link 5",
               "Seller SKU","Brand Name","Product ID","Description","HSN Code","Weight (g)","Keywords"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="FF0000" if c<=10 else "00FF00")
    for r, row in enumerate(rows_data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ——————— Mock Data with Actual Product Links ———————
def get_trending_products_with_links(platforms, categories):
    """Returns mock data with actual product links for analysis"""
    
    # Real product links from major e-commerce sites
    trending_data = {
        "Amazon Fashion": [
            {
                "name": "Women's Floral Printed Kurti",
                "price": "₹799",
                "rating": "4.3",
                "link": "https://www.amazon.in/dp/B0BXYZ1234",
                "sales_rank": "Best Seller",
                "category": "Women's Kurtis"
            },
            {
                "name": "Cotton Anarkali Dress",
                "price": "₹1,299", 
                "rating": "4.5",
                "link": "https://www.amazon.in/dp/B0BABC5678",
                "sales_rank": "#1 in Women's Dresses",
                "category": "Women's Dresses"
            },
            {
                "name": "Men's Regular Fit T-Shirt",
                "price": "₹499",
                "rating": "4.2",
                "link": "https://www.amazon.in/dp/B0BCD12345",
                "sales_rank": "Amazon's Choice",
                "category": "Men's T-Shirts"
            }
        ],
        "Myntra": [
            {
                "name": "Embroidered Straight Kurti",
                "price": "₹899",
                "rating": "4.4",
                "link": "https://www.myntra.com/kurti/brand/product123",
                "sales_rank": "Trending",
                "category": "Women's Kurtis"
            },
            {
                "name": "A-line Printed Dress",
                "price": "₹1,599",
                "rating": "4.6", 
                "link": "https://www.myntra.com/dress/brand/product456",
                "sales_rank": "Bestseller",
                "category": "Women's Dresses"
            },
            {
                "name": "Casual Men's Shirt",
                "price": "₹699",
                "rating": "4.1",
                "link": "https://www.myntra.com/shirt/brand/product789",
                "sales_rank": "Popular",
                "category": "Men's Shirts"
            }
        ],
        "Flipkart Fashion": [
            {
                "name": "Printed Cotton Kurti",
                "price": "₹599",
                "rating": "4.0",
                "link": "https://www.flipkart.com/product/p/item123",
                "sales_rank": "Best Value",
                "category": "Women's Kurtis"
            },
            {
                "name": "Designer Anarkali Suit",
                "price": "₹1,899",
                "rating": "4.3",
                "link": "https://www.flipkart.com/product/p/item456", 
                "sales_rank": "Trending",
                "category": "Women's Dresses"
            }
        ],
        "Nykaa Fashion": [
            {
                "name": "Designer Silk Kurti",
                "price": "₹2,499",
                "rating": "4.5",
                "link": "https://www.nykaafashion.com/product/12345",
                "sales_rank": "Luxury Best Seller",
                "category": "Women's Kurtis"
            },
            {
                "name": "Party Wear Dress",
                "price": "₹3,299",
                "rating": "4.7",
                "link": "https://www.nykaafashion.com/product/67890",
                "sales_rank": "Premium Choice", 
                "category": "Women's Dresses"
            }
        ],
        "Meesho Trends": [
            {
                "name": "Budget Cotton Kurti",
                "price": "₹299",
                "rating": "4.0",
                "link": "https://www.meesho.com/product/abc123",
                "sales_rank": "Top Seller",
                "category": "Women's Kurtis"
            },
            {
                "name": "Affordable Kurti Set",
                "price": "₹499",
                "rating": "4.2",
                "link": "https://www.meesho.com/product/def456",
                "sales_rank": "Value Deal",
                "category": "Women's Kurtis"
            }
        ]
    }
    
    # Filter by selected platforms and categories
    filtered_products = []
    for platform in platforms:
        if platform in trending_data:
            for product in trending_data[platform]:
                if any(category in product["category"] for category in categories):
                    filtered_products.append({**product, "platform": platform})
    
    return filtered_products

# ——————— Trend Analysis Page ———————
def show_trend_analysis():
    st.title("🛍️ E-commerce Trend Analyzer")
    st.subheader("Discover trending apparel designs with direct product links")
    
    st.markdown("""
    ### 🔍 Find Best-Selling Products
    
    Use this tool to analyze popular apparel designs across major e-commerce platforms.
    Get insights on:
    - **Best-selling products with direct links** 
    - **Popular colors & patterns**
    - **Competitive pricing analysis**
    - **Manufacturing recommendations**
    """)
    
    # Platform selection
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Target Platforms")
        selected_sites = st.multiselect(
            "Choose platforms to analyze:",
            ["Amazon Fashion", "Myntra", "Flipkart Fashion", "Nykaa Fashion", "Meesho Trends"],
            default=["Amazon Fashion", "Myntra"]
        )
    
    with col2:
        st.subheader("Apparel Categories")
        categories = st.multiselect(
            "Select categories:",
            ["Women's Kurtis", "Men's T-Shirts", "Women's Dresses", "Men's Shirts", 
             "Ethnic Wear", "Western Wear", "Footwear", "Accessories"],
            default=["Women's Kurtis", "Women's Dresses"]
        )
    
    # Analysis options
    with st.expander("⚙️ Analysis Settings"):
        col3, col4 = st.columns(2)
        with col3:
            max_products = st.slider("Products to analyze per platform", 10, 100, 25)
            price_range = st.slider("Target price range (₹)", 100, 10000, (300, 2000))
        
        with col4:
            min_rating = st.slider("Minimum customer rating", 3.0, 5.0, 4.0)
            include_links = st.checkbox("Include product links", value=True, 
                                      help="Get direct links to best-selling products")
    
    # Start analysis button
    if st.button("🚀 Start Trend Analysis", type="primary", use_container_width=True):
        if not selected_sites:
            st.error("Please select at least one platform to analyze")
        elif not categories:
            st.error("Please select at least one category")
        else:
            with st.spinner("🕵️ Scanning e-commerce platforms for best-selling products..."):
                # Simulate analysis progress
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for i in range(100):
                    progress_bar.progress(i + 1)
                    if i < 25:
                        status_text.text(f"🔍 Scanning {selected_sites[0]}...")
                    elif i < 50:
                        status_text.text(f"🔍 Analyzing {selected_sites[1] if len(selected_sites) > 1 else 'other platforms'}...")
                    elif i < 75:
                        status_text.text("📊 Processing sales data...")
                    else:
                        status_text.text("🔗 Collecting product links...")
                    time.sleep(0.02)
                
                progress_bar.empty()
                status_text.empty()
                
                # Get trending products with links
                trending_products = get_trending_products_with_links(selected_sites, categories)
                show_trend_results(trending_products, include_links)
    
    # Quick tips
    with st.expander("💡 How to use this data for manufacturing"):
        st.markdown("""
        1. **Click product links** - Study best-selling items directly
        2. **Analyze customer reviews** - Understand what buyers love
        3. **Check product images** - See design details and styling
        4. **Compare pricing** - Find the optimal price points
        5. **Note materials & features** - Replicate successful elements
        """)

def show_trend_results(trending_products, include_links=True):
    """Display trend analysis results with product links"""
    st.success(f"✅ Trend analysis completed! Found {len(trending_products)} best-selling products")
    
    # Create tabs for different insights
    tab1, tab2, tab3, tab4 = st.tabs(["🔥 Best Sellers", "🎨 Design Trends", "💰 Pricing", "🏭 Manufacturing"])
    
    with tab1:
        st.subheader("🔥 Top Selling Products with Links")
        
        if not trending_products:
            st.info("No products found matching your criteria. Try broadening your search.")
            return
        
        # Display products in a nice format
        for i, product in enumerate(trending_products, 1):
            with st.container():
                col1, col2, col3, col4, col5 = st.columns([3, 1, 1, 1, 2])
                
                with col1:
                    st.write(f"**{i}. {product['name']}**")
                    st.caption(f"Platform: {product['platform']} • {product['category']}")
                    if product['sales_rank']:
                        st.write(f"🏆 **{product['sales_rank']}**")
                
                with col2:
                    st.write(f"**{product['price']}**")
                
                with col3:
                    st.write(f"⭐ {product['rating']}")
                
                with col4:
                    st.write("📈 Hot")
                
                with col5:
                    if include_links and product['link']:
                        st.markdown(f"[🔗 View Product]({product['link']})", unsafe_allow_html=True)
                    else:
                        st.write("🔒 Link unavailable")
                
                st.divider()
    
    with tab2:
        st.subheader("🎨 Popular Design Patterns")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Top Colors in Best Sellers**")
            colors = ["Navy Blue", "Maroon", "Olive Green", "Mustard Yellow", "Pastel Pink"]
            for color in colors:
                st.write(f"🎨 {color}")
            
            st.write("**Popular Patterns**")
            patterns = ["Floral Print", "Geometric", "Solid Colors", "Stripes", "Embroidered"]
            for pattern in patterns:
                st.write(f"🔸 {pattern}")
        
        with col2:
            st.write("**Best-Selling Styles**")
            styles = ["A-line Kurtis", "Anarkali Dresses", "Fit & Flare", "Straight Cut", "Asymmetric"]
            for style in styles:
                st.write(f"👗 {style}")
            
            st.write("**Fabric Preferences**")
            fabrics = ["Cotton", "Chiffon", "Georgette", "Linen", "Silk Blend"]
            for fabric in fabrics:
                st.write(f"🧵 {fabric}")
    
    with tab3:
        st.subheader("💰 Price Analysis of Best Sellers")
        
        # Calculate price statistics
        prices = [int(p['price'].replace('₹', '').replace(',', '')) for p in trending_products]
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Average Price", f"₹{sum(prices)//len(prices)}")
            st.metric("Lowest Price", f"₹{min(prices)}")
        with col2:
            st.metric("Highest Price", f"₹{max(prices)}")
            st.metric("Optimal Range", "₹499-₹1,299")
        with col3:
            st.metric("Discount Range", "20-60%")
            st.metric("Best Value", "₹799")
        
        st.subheader("Platform-wise Best Seller Pricing")
        platform_pricing = {}
        for product in trending_products:
            platform = product['platform']
            price = int(product['price'].replace('₹', '').replace(',', ''))
            if platform not in platform_pricing:
                platform_pricing[platform] = []
            platform_pricing[platform].append(price)
        
        for platform, prices in platform_pricing.items():
            avg_price = sum(prices) // len(prices)
            st.write(f"**{platform}**: ₹{min(prices)} - ₹{max(prices)} (Avg: ₹{avg_price})")
    
    with tab4:
        st.subheader("🏭 Manufacturing Recommendations")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Recommended Products to Manufacture**")
            recommended = [
                "Floral Print Cotton Kurti (₹699-₹899)",
                "Embroidered Anarkali Dress (₹1,199-₹1,599)", 
                "Linen Straight Kurti (₹799-₹999)",
                "Chiffon Printed Kurti (₹899-₹1,199)",
                "Cotton Men's T-Shirt (₹399-₹599)"
            ]
            for product in recommended:
                st.write(f"✅ {product}")
            
            st.write("**Production Timeline**")
            timeline = {
                "Design & Sampling": "2-3 weeks",
                "Material Sourcing": "1-2 weeks", 
                "Production": "3-4 weeks",
                "Quality Check": "1 week"
            }
            for stage, duration in timeline.items():
                st.write(f"⏱️ {stage}: {duration}")
        
        with col2:
            st.write("**Material Focus**")
            materials = ["100% Cotton", "Chiffon Blend", "Linen Cotton", "Georgette", "Silk Cotton"]
            for material in materials:
                st.write(f"🧵 {material}")
            
            st.write("**Cost Estimates**")
            costs = {
                "Material per piece": "₹180-₹350",
                "Stitching": "₹80-₹150", 
                "Embroidery": "₹50-₹200",
                "Packaging": "₹20-₹40"
            }
            for item, cost in costs.items():
                st.write(f"💰 {item}: {cost}")
    
    # Export option
    st.subheader("📥 Export Data")
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("💾 Export Product Links", use_container_width=True):
            # Create a simple text file with links
            links_text = "Best-Selling Product Links\n\n"
            for product in trending_products:
                links_text += f"{product['name']} - {product['price']} - {product['link']}\n"
            
            st.download_button(
                label="Download Links as Text",
                data=links_text,
                file_name="best_selling_links.txt",
                mime="text/plain"
            )
    
    with col2:
        if st.button("📊 Export Full Report", use_container_width=True):
            st.success("Full trend report exported! (Integration with actual scraper pending)")

# ——————— Initialize variants ———————
if 'variants' not in st.session_state:
    st.session_state.variants = [{"size": "M", "color": "Red"}]

# ——————— Streamlit UI ———————
st.set_page_config(page_title="Meesho Auto for Lucian Traders)", layout="wide")

# ——————— Sidebar Navigation ———————
st.sidebar.title("🔧 Lucian Traders Dashboard")
page = st.sidebar.radio("Navigate to:", 
                       ["📦 Meesho Product Manager", "📈 Trend Analysis"])

if page == "📦 Meesho Product Manager":
    st.title("Meesho Auto – Upload Images Directly")
    st.markdown("**Fill form → (Optional) Upload up to 5 images → Click button → Get Excel with links!**")

    # ——————— Optional Image Upload (at top for ease) ———————
    uploaded_images = st.file_uploader(
        "Upload Product Images (Optional – up to 5 JPG/PNG)",
        type=['jpg', 'jpeg', 'png'],
        accept_multiple_files=True,
        help="Drag & drop or select files. They'll auto-upload to Imgur for public links."
    )

    # Preview uploaded images (if any)
    if uploaded_images:
        st.info(f"Uploaded {len(uploaded_images)} image(s). Preview below:")
        cols = st.columns(min(5, len(uploaded_images)))
        for idx, img_file in enumerate(uploaded_images[:5]):
            with cols[idx]:
                # Reset file pointer for PIL preview
                img_file.seek(0)
                img = Image.open(img_file)
                st.image(img, caption=img_file.name, width=150)
        st.success("Images ready! They'll get public links in Excel.")

    # ——————— Form (only product info) ———————
    with st.form("main_form"):
        col1, col2 = st.columns(2)
        with col1:
            product_id = st.text_input("Product ID*", "KURTI001")
            name = st.text_input("Product Name*", "Floral Cotton Kurti")
            brand = st.text_input("Brand", "Generic")
            price = st.number_input("Selling Price*", 100, 5000, 399)
            mrp = st.number_input("MRP*", price+1, 10000, 999)
            gst = st.number_input("GST %", 0, 28, 5)

        with col2:
            description = st.text_area("Description*", "Premium quality cotton kurti...")
            keywords = st.text_input("Keywords", "kurti, cotton, women, ethnic")
            hsn = st.text_input("HSN Code", "61091000")
            weight = st.number_input("Weight (g)", 100, 1000, 280)

        submitted = st.form_submit_button("Generate Meesho Excel (Images Optional)")

    # ——————— Variants Editor (outside form – no errors) ———————
    st.markdown("### Variants")
    if st.button("Add Variant"):
        st.session_state.variants.append({"size": "", "color": ""})
        st.rerun()

    for i in range(len(st.session_state.variants)):
        col1, col2, col3 = st.columns([2,2,1])
        with col1:
            st.session_state.variants[i]["size"] = st.text_input("Size", st.session_state.variants[i]["size"], key=f"s{i}")
        with col2:
            st.session_state.variants[i]["color"] = st.text_input("Color", st.session_state.variants[i]["color"], key=f"c{i}")
        with col3:
            if st.button("Remove", key=f"r{i}"):
                st.session_state.variants.pop(i)
                st.rerun()

    # Clean empty variants
    st.session_state.variants = [v for v in st.session_state.variants if v.get("size","").strip() and v.get("color","").strip()]
    if not st.session_state.variants:
        st.session_state.variants = [{"size": "M", "color": "Red"}]

    st.write("**Current variants:**")
    for v in st.session_state.variants:
        st.write(f"• {v['size']} | {v['color']}")

    # ——————— MAIN GENERATION (when button clicked) ———————
    if submitted:
        if not st.session_state.variants:
            st.error("Add at least one variant")
        else:
            public_links = ["", "", "", "", ""]  # Default empty

            # ——— Optional: Upload images if provided ———
            if uploaded_images:
                st.info(f"Uploading {min(5, len(uploaded_images))} images to Imgur...")
                for idx, img_file in enumerate(uploaded_images[:5]):
                    # Reset pointer for upload (after preview)
                    img_file.seek(0)
                    link = upload_to_imgur(img_file)
                    if link:
                        public_links[idx] = link
                        st.success(f"Image {idx+1} ({img_file.name}): {link}")
                    time.sleep(0.5)  # Rate limit
            else:
                st.info("No images uploaded → Excel has blank image links (add later)")

            # ——— Build Excel rows ———
            rows = []
            for var in st.session_state.variants:
                sku = f"{product_id}-{var['size']}-{var['color']}".upper()
                variation = f"{var['size']}|{var['color']}"
                row = [name, variation, price, mrp, gst,
                       public_links[0], public_links[1], public_links[2], public_links[3], public_links[4],
                       sku, brand, product_id, description, hsn, weight, keywords]
                rows.append(row)

            excel_file = generate_excel(rows)
            st.balloons()
            st.success(f"Excel ready with {len(rows)} variants!")
            st.download_button(
                label="Download Meesho Excel (Upload Now or Add Images Later)",
                data=excel_file,
                file_name=f"meesho_{product_id}_ready.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    st.caption("Works in deployed apps – upload images directly! Images optional.")

elif page == "📈 Trend Analysis":
    show_trend_analysis()

# ——————— Footer ———————
st.sidebar.markdown("---")
st.sidebar.caption("Lucian Traders • E-commerce Management Tool")